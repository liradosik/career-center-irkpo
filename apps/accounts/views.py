from datetime import date, timedelta
import base64
import csv
import io
import uuid
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Count, F, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_GET

from apps.courses.models import Course, CourseRegistration, StudentFavoriteCourse
from apps.portfolio.models import PortfolioEntry
from apps.vacancies.models import StudentFavoriteVacancy, Vacancy, VacancyResponse

from .decorators import role_required
from .forms import (
    AdminCourseForm,
    AdminCuratorCreateForm,
    AdminCuratorUpdateForm,
    AdminSpecialtyForm,
    AdminStudentCreateForm,
    AdminStudyGroupForm,
    AdminStudentUpdateForm,
    AdminVacancyForm,
    CuratorStudentAcademicStatusForm,
    CuratorImportForm,
    EmailAuthenticationForm,
    GroupImportForm,
    AccountPasswordChangeForm,
    StudentImportForm,
    SupportTicketAdminUpdateForm,
    SupportTicketCreateForm,
    PublicSupportTicketCreateForm,
    StudentProfileForm,
    UserStudentForm,
    StudentAcademicReadonlyForm,
    UserProfileSettingsForm,
    sync_student_with_group,
)
from .models import ActivityLog, AdminActivityLog, Specialty, StudyGroup, SupportTicket, User
from .utils import apply_user_photo_update, log_admin_action


class CustomLoginView(LoginView):
    template_name = 'accounts/login.html'
    authentication_form = EmailAuthenticationForm


class CustomLogoutView(LogoutView):
    pass


IMPORT_TEMPLATES = {
    'students': {
        'sheet_title': 'Студенты',
        'headers': ['full_name', 'email', 'password', 'group'],
        'example': ['Иванова Анна Сергеевна', 'anna.ivanova@example.com', 'TempPass123', 'Н-121/2'],
        'download_filename': 'шаблон_импорта_студентов',
    },
    'curators': {
        'sheet_title': 'Кураторы',
        'headers': ['full_name', 'email', 'password'],
        'example': ['Иванова Ольга Сергеевна', 'ivanova.curator@example.com', 'TempPass123'],
        'download_filename': 'шаблон_импорта_кураторов',
    },
    'groups': {
        'sheet_title': 'Группы',
        'headers': ['specialty_letter', 'admission_year', 'course_number', 'subgroup_number', 'curator_email'],
        'example': ['Н', '2021', '1', '2', 'ivanova.curator@example.com'],
        'download_filename': 'шаблон_импорта_групп',
    },
}

IMPORT_VALIDATION_CONFIG = {
    'students': {
        'required_columns': ['full_name', 'email', 'password', 'group'],
        'display_name': 'студентов',
    },
    'curators': {
        'required_columns': ['full_name', 'email', 'password'],
        'display_name': 'кураторов',
    },
    'groups': {
        'required_columns': ['specialty_letter', 'admission_year', 'course_number', 'subgroup_number'],
        'display_name': 'групп',
    },
    'specialties': {
        'required_columns': ['specialty_code', 'specialty_letter', 'name'],
        'display_name': 'специальностей',
    },
    'vacancies': {
        'required_columns': ['title', 'company', 'employment_type', 'format', 'direction', 'contacts', 'status'],
        'display_name': 'вакансий',
    },
    'courses': {
        'required_columns': ['title', 'organization', 'type', 'format', 'date', 'seats', 'status'],
        'display_name': 'курсов',
    },
}

IMPORT_TYPE_HINTS = [
    ({'email', 'full_name', 'password', 'group'}, 'студентов'),
    ({'email', 'full_name', 'password'}, 'кураторов'),
    ({'company', 'employment_type'}, 'вакансий'),
    ({'organization', 'date', 'seats'}, 'курсов'),
]


def _normalize_header_name(value):
    return ' '.join(str(value or '').strip().lower().split())


def _build_import_file_type_hint(headers, expected_import_type):
    for columns, import_name in IMPORT_TYPE_HINTS:
        if columns.issubset(headers):
            if expected_import_type == 'students' and import_name in {'студентов', 'кураторов'}:
                continue
            if expected_import_type == 'curators' and import_name in {'студентов', 'кураторов'}:
                continue
            return import_name
    return None


def parse_import_file(uploaded_file, import_type):
    validation_config = IMPORT_VALIDATION_CONFIG[import_type]
    required_columns = validation_config['required_columns']
    import_display_name = validation_config['display_name']

    file_name = (uploaded_file.name or '').lower()
    if file_name.endswith('.csv'):
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        fieldnames = reader.fieldnames or []
        rows = list(reader)
    elif file_name.endswith('.xlsx'):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            fieldnames = []
            rows = []
        else:
            fieldnames = [str(value).strip() if value is not None else '' for value in all_rows[0]]
            rows = []
            for raw_row in all_rows[1:]:
                row_map = {}
                for idx, key in enumerate(fieldnames):
                    value = raw_row[idx] if idx < len(raw_row) else ''
                    row_map[key] = '' if value is None else str(value).strip()
                rows.append(row_map)
    else:
        return [], ['Поддерживаются только CSV и XLSX']

    normalized_headers = {_normalize_header_name(header) for header in fieldnames if _normalize_header_name(header)}
    if not normalized_headers:
        return [], ['Файл пустой или не содержит строку с заголовками.']

    missing = [col for col in required_columns if col not in normalized_headers]
    if missing:
        expected_columns = ', '.join(required_columns)
        hint_import_name = _build_import_file_type_hint(normalized_headers, import_type)
        if hint_import_name:
            return [], [
                f'Похоже, загружен файл {hint_import_name}. Для импорта {import_display_name} нужен файл с колонками: {expected_columns}.'
            ]
        return [], [
            f'Файл не соответствует шаблону импорта {import_display_name}. Отсутствуют колонки: {", ".join(missing)}.'
        ]

    header_map = {_normalize_header_name(header): header for header in fieldnames if _normalize_header_name(header)}
    normalized_rows = []
    for row in rows:
        normalized_rows.append({col: (row.get(header_map[col]) or '').strip() for col in normalized_headers})
    return normalized_rows, []


def set_download_filename(response, base_filename, extension):
    filename = f'{base_filename}.{extension}'
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"


def is_empty_import_row(row):
    return not any((value or '').strip() for value in row.values())


def normalize_group_code(raw_value):
    value = (raw_value or '').strip()
    value = ' '.join(value.split())
    value = value.replace(' - ', '-').replace(' -', '-').replace('- ', '-')
    value = value.replace(' / ', '/').replace(' /', '/').replace('/ ', '/')
    return value


def build_template_csv_response(template_key):
    template_data = IMPORT_TEMPLATES[template_key]
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    set_download_filename(response, template_data['download_filename'], 'csv')
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(template_data['headers'])
    writer.writerow(template_data['example'])
    return response


def build_template_xlsx_response(template_key):
    template_data = IMPORT_TEMPLATES[template_key]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = template_data['sheet_title']

    sheet.append(template_data['headers'])
    sheet.append(template_data['example'])

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for idx, header in enumerate(template_data['headers'], start=1):
        max_len = max(len(str(header)), len(str(template_data['example'][idx - 1])))
        sheet.column_dimensions[get_column_letter(idx)].width = max_len + 6

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    set_download_filename(response, template_data['download_filename'], 'xlsx')
    workbook.save(response)
    return response


def role_redirect(user):
    if user.role == User.Role.STUDENT:
        return reverse('accounts:student_dashboard')
    if user.role == User.Role.CURATOR:
        return reverse('accounts:curator_dashboard')
    return reverse('accounts:admin_dashboard')


def sync_group_students(group):
    students = User.objects.filter(role=User.Role.STUDENT, study_group=group)
    for student in students:
        sync_student_with_group(student, group)
        student.save(update_fields=['study_group', 'group', 'specialty', 'admission_year', 'curator'])


def curator_students_queryset(curator, include_graduates=False):
    qs = User.objects.filter(role=User.Role.STUDENT).filter(
        Q(study_group__curator=curator) | Q(curator=curator)
    )
    if include_graduates:
        return qs.distinct()

    qs = qs.filter(study_group__is_active=True).exclude(academic_status=User.AcademicStatus.GRADUATED)
    return qs.distinct()


@require_GET
def home(request):
    return render(request, 'public/index.html')


def redirect_by_role(request):
    return redirect(role_redirect(request.user))


@login_required
def change_password(request):
    forced = bool(request.user.must_change_password)
    if request.method == 'POST':
        form = AccountPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            user.must_change_password = False
            user.save(update_fields=['password', 'must_change_password'])
            update_session_auth_hash(request, user)
            messages.success(request, 'Пароль успешно изменён.')
            return redirect('accounts:redirect_by_role')
    else:
        form = AccountPasswordChangeForm(request.user)
    return render(request, 'accounts/change_password.html', {'form': form, 'forced': forced})


@role_required(User.Role.STUDENT)
def student_dashboard(request):
    entries_qs = PortfolioEntry.objects.filter(student=request.user)
    recent_entries = entries_qs.order_by('-created_at')[:5]

    registrations_qs = (
        CourseRegistration.objects
        .filter(student=request.user)
        .select_related('course')
        .order_by('-created_at')
    )
    registrations = registrations_qs[:5]

    resume = getattr(request.user, 'resume_settings', None)
    profile = getattr(request.user, 'student_profile', None)
    resume_public_url = ''
    if profile:
        resume_public_url = request.build_absolute_uri(f"/resumes/public/{profile.public_resume_token}/")

    current_course = None

    if request.user.study_group and request.user.study_group.course_number:
        current_course = request.user.study_group.course_number
    elif request.user.admission_year:
        current_course = max(date.today().year - request.user.admission_year + 1, 1)

    curator = None
    if request.user.study_group and request.user.study_group.curator:
        curator = request.user.study_group.curator
    elif request.user.curator:
        curator = request.user.curator

    context = {
        'recent_entries': recent_entries,
        'portfolio_total': entries_qs.count(),
        'portfolio_pending': entries_qs.filter(status=PortfolioEntry.Status.PENDING).count(),
        'portfolio_approved': entries_qs.filter(status=PortfolioEntry.Status.APPROVED).count(),
        'portfolio_rejected': entries_qs.filter(status=PortfolioEntry.Status.REJECTED).count(),
        'resume': resume,
        'resume_is_public': bool(resume and resume.is_public and profile),
        'resume_public_url': resume_public_url,
        'resume_updated_at': getattr(resume, "updated_at", None) if resume else None,
        'registrations': registrations,
        'registered_courses_count': registrations_qs.filter(status=CourseRegistration.Status.REGISTERED).count(),
        'current_course': current_course,
        'profile_incomplete': not all([request.user.group, request.user.specialty, request.user.admission_year]),
        'curator': curator,
    }
    return render(request, 'dashboard/student_dashboard.html', context)


def _save_account_support_ticket(form, user):
    ticket = form.save(commit=False)
    ticket.requester = user
    ticket.student = user if user.role == User.Role.STUDENT else None
    ticket.requester_type = user.role if user.role in {User.Role.STUDENT, User.Role.CURATOR} else SupportTicket.RequesterType.UNKNOWN
    ticket.source = SupportTicket.Source.ACCOUNT
    ticket.save()
    return ticket


@role_required(User.Role.STUDENT)
def student_support_tickets(request):
    tickets = SupportTicket.objects.filter(requester=request.user).order_by('-created_at')
    form = SupportTicketCreateForm()
    if request.method == 'POST':
        form = SupportTicketCreateForm(request.POST)
        if form.is_valid():
            _save_account_support_ticket(form, request.user)
            messages.success(request, 'Обращение отправлено в техподдержку.')
            return redirect('accounts:student_support_tickets')
    return render(request, 'support/student_tickets.html', {'tickets': tickets, 'form': form, 'support_title': 'Техподдержка'})



@login_required
def support_knowledge(request):
    knowledge_categories = [
        {
            'name': 'Вход и аккаунт',
            'icon': 'bi-person-lock',
            'items': [
                {
                    'question': 'Не получается войти в аккаунт',
                    'answer': 'Проверьте email и пароль. Если пароль не подходит или аккаунт не найден, создайте обращение в техподдержку.',
                    'icon': 'bi-box-arrow-in-right',
                },
                {
                    'question': 'Я забыл пароль',
                    'answer': 'Обратитесь в техподдержку и укажите ФИО, группу и email, который использовался для входа.',
                    'icon': 'bi-key',
                },
                {
                    'question': 'В данных профиля ошибка',
                    'answer': 'Студент может изменить контакты и описание, но группа, специальность и учебный статус редактируются администрацией.',
                    'icon': 'bi-exclamation-diamond',
                },
            ],
        },
        {'name': 'Профиль', 'icon': 'bi-person-vcard', 'items': []},
        {'name': 'Портфолио', 'icon': 'bi-collection', 'items': [
            {
                'question': 'Почему работа не отображается в резюме?',
                'answer': 'В резюме попадают только подтверждённые записи портфолио.',
                'icon': 'bi-journal-check',
            },
            {
                'question': 'Что значит статус «Ожидает проверки»?',
                'answer': 'Запись отправлена куратору. После проверки она будет подтверждена или отклонена.',
                'icon': 'bi-hourglass-split',
            },
            {
                'question': 'Можно ли прикреплять файлы?',
                'answer': 'Да, к работе портфолио можно прикрепить файл или ссылку, если это предусмотрено формой.',
                'icon': 'bi-paperclip',
            },
        ]},
        {'name': 'Резюме', 'icon': 'bi-file-earmark-person', 'items': [
            {
                'question': 'Как открыть публичное резюме?',
                'answer': 'В конструкторе включите публичное резюме, сохраните настройки и нажмите «Открыть публичное резюме».',
                'icon': 'bi-globe2',
            },
            {
                'question': 'Почему публичное резюме недоступно?',
                'answer': 'Проверьте, включён ли публичный доступ в конструкторе резюме.',
                'icon': 'bi-eye-slash',
            },
            {
                'question': 'Почему PDF выглядит не так?',
                'answer': 'Сначала выберите стиль резюме, нажмите «Сохранить», затем скачайте PDF.',
                'icon': 'bi-filetype-pdf',
            },
        ]},
        {'name': 'Курсы', 'icon': 'bi-mortarboard', 'items': [
            {
                'question': 'Как записаться на курс?',
                'answer': 'Откройте раздел «Курсы и семинары» и нажмите кнопку записи на подходящем курсе.',
                'icon': 'bi-bookmark-plus',
            },
            {
                'question': 'Можно ли отменить запись?',
                'answer': 'Да, если вы уже записаны, в карточке курса появится кнопка отмены.',
                'icon': 'bi-bookmark-dash',
            },
            {
                'question': 'Почему нет кнопки записи?',
                'answer': 'Курс может быть скрыт, архивирован, недоступен или на очном курсе закончились места.',
                'icon': 'bi-slash-circle',
            },
        ]},
        {'name': 'Вакансии', 'icon': 'bi-briefcase', 'items': [
            {
                'question': 'Как откликнуться на вакансию?',
                'answer': 'Откройте раздел «Вакансии», выберите предложение и нажмите «Откликнуться».',
                'icon': 'bi-send-check',
            },
            {
                'question': 'Можно ли сохранить вакансию?',
                'answer': 'Да, нажмите на значок закладки. Сохранённые вакансии находятся в разделе «Избранное».',
                'icon': 'bi-bookmark-heart',
            },
            {
                'question': 'Где посмотреть избранное?',
                'answer': 'В левом меню студента есть раздел «Избранное».',
                'icon': 'bi-stars',
            },
        ]},
    ]

    return render(request, 'accounts/support_knowledge.html', {'knowledge_categories': knowledge_categories})


@role_required(User.Role.STUDENT)
def student_favorites(request):
    favorite_vacancies = (
        StudentFavoriteVacancy.objects
        .filter(student=request.user)
        .select_related('vacancy')
        .order_by('-created_at')
    )
    favorite_courses = (
        StudentFavoriteCourse.objects
        .filter(student=request.user)
        .select_related('course')
        .order_by('-created_at')
    )

    course_ids = favorite_courses.values_list('course_id', flat=True)
    course_registrations = CourseRegistration.objects.filter(
        student=request.user,
        course_id__in=course_ids,
    )
    registration_map = {registration.course_id: registration for registration in course_registrations}

    responded_vacancy_ids = set(
        VacancyResponse.objects
        .filter(student=request.user)
        .values_list('vacancy_id', flat=True)
    )

    return render(request, 'accounts/student_favorites.html', {
        'favorite_vacancies': favorite_vacancies,
        'favorite_courses': favorite_courses,
        'registration_map': registration_map,
        'responded_vacancy_ids': responded_vacancy_ids,
    })


@role_required(User.Role.CURATOR)
def curator_support_tickets(request):
    tickets = SupportTicket.objects.filter(requester=request.user).order_by('-created_at')
    form = SupportTicketCreateForm()
    if request.method == 'POST':
        form = SupportTicketCreateForm(request.POST)
        if form.is_valid():
            _save_account_support_ticket(form, request.user)
            messages.success(request, 'Обращение отправлено в техподдержку.')
            return redirect('accounts:curator_support_tickets')
    return render(request, 'support/student_tickets.html', {'tickets': tickets, 'form': form, 'support_title': 'Техподдержка куратора'})


def public_support_ticket_create(request):
    if request.user.is_authenticated:
        if request.user.role == User.Role.STUDENT:
            return redirect('accounts:student_support_tickets')
        if request.user.role == User.Role.CURATOR:
            return redirect('accounts:curator_support_tickets')
        return redirect('accounts:admin_support_tickets')

    form = PublicSupportTicketCreateForm()
    if request.method == 'POST':
        form = PublicSupportTicketCreateForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.source = SupportTicket.Source.PUBLIC
            ticket.requester_type = form.cleaned_data['requester_type']
            ticket.save()
            messages.success(request, 'Обращение отправлено. Администратор свяжется с вами по указанным контактам.', extra_tags='public-support')
            return redirect('accounts:public_support_ticket_create')
    return render(request, 'support/public_ticket_form.html', {'form': form})


@role_required(User.Role.CURATOR)
def curator_dashboard(request):
    students = curator_students_queryset(request.user)
    all_students = curator_students_queryset(request.user, include_graduates=True)
    student_ids = students.values_list('id', flat=True)

    pending_entries_qs = (
        PortfolioEntry.objects
        .filter(student_id__in=student_ids, status=PortfolioEntry.Status.PENDING)
        .select_related('student')
        .order_by('-created_at')
    )

    recent_activity = ActivityLog.objects.filter(student_id__in=student_ids).select_related('student')[:7]

    context = {
        'students_count': students.count(),
        'studying_count': students.filter(academic_status=User.AcademicStatus.STUDYING).count(),
        'academic_leave_count': students.filter(academic_status=User.AcademicStatus.ACADEMIC_LEAVE).count(),
        'expelled_count': students.filter(academic_status=User.AcademicStatus.EXPELLED).count(),
        'graduated_count': all_students.filter(academic_status=User.AcademicStatus.GRADUATED).count(),
        'pending_count': pending_entries_qs.count(),
        'approved_count': PortfolioEntry.objects.filter(
            student_id__in=student_ids, status=PortfolioEntry.Status.APPROVED
        ).count(),
        'rejected_count': PortfolioEntry.objects.filter(
            student_id__in=student_ids, status=PortfolioEntry.Status.REJECTED
        ).count(),
        'pending_entries': pending_entries_qs[:5],
        'recent_activity': recent_activity,
    }
    return render(request, 'curator/dashboard.html', context)


@role_required(User.Role.CURATOR)
def curator_activity(request):
    period_options = {
        '30d': {'label': 'За последние 30 дней', 'days': 30},
        '6m': {'label': 'За последние 6 месяцев', 'days': 182},
        '1y': {'label': 'За последний год', 'days': 365},
    }
    selected_period = request.GET.get('period', '30d')
    if selected_period not in period_options:
        selected_period = '30d'

    period_days = period_options[selected_period]['days']
    now = timezone.now()
    period_start = now - timedelta(days=period_days)
    prev_period_start = period_start - timedelta(days=period_days)

    include_graduates = request.GET.get('graduates') == '1'
    students = curator_students_queryset(request.user, include_graduates=include_graduates)
    student_ids = students.values_list('id', flat=True)

    current_registrations = CourseRegistration.objects.filter(
        student_id__in=student_ids,
        status=CourseRegistration.Status.REGISTERED,
        created_at__gte=period_start,
        created_at__lt=now,
    )
    current_responses = VacancyResponse.objects.filter(
        student_id__in=student_ids,
        created_at__gte=period_start,
        created_at__lt=now,
    )
    current_portfolio = PortfolioEntry.objects.filter(
        student_id__in=student_ids,
        created_at__gte=period_start,
        created_at__lt=now,
    )

    prev_registrations = CourseRegistration.objects.filter(
        student_id__in=student_ids,
        status=CourseRegistration.Status.REGISTERED,
        created_at__gte=prev_period_start,
        created_at__lt=period_start,
    )
    prev_responses = VacancyResponse.objects.filter(
        student_id__in=student_ids,
        created_at__gte=prev_period_start,
        created_at__lt=period_start,
    )
    prev_portfolio = PortfolioEntry.objects.filter(
        student_id__in=student_ids,
        created_at__gte=prev_period_start,
        created_at__lt=period_start,
    )

    registrations_total = current_registrations.count()
    responses_total = current_responses.count()
    portfolio_total = current_portfolio.count()

    active_student_ids = set(current_registrations.values_list('student_id', flat=True))
    active_student_ids.update(current_responses.values_list('student_id', flat=True))
    active_student_ids.update(current_portfolio.values_list('student_id', flat=True))
    active_students_total = len(active_student_ids)

    prev_active_student_ids = set(prev_registrations.values_list('student_id', flat=True))
    prev_active_student_ids.update(prev_responses.values_list('student_id', flat=True))
    prev_active_student_ids.update(prev_portfolio.values_list('student_id', flat=True))
    prev_active_students_total = len(prev_active_student_ids)

    def calculate_trend(current, previous):
        if previous == 0:
            if current == 0:
                return 0, 'same'
            return 100, 'up'

        trend_percent = int(round(((current - previous) / previous) * 100))
        if trend_percent > 0:
            return trend_percent, 'up'
        if trend_percent < 0:
            return abs(trend_percent), 'down'
        return 0, 'same'

    registrations_trend_percent, registrations_trend_direction = calculate_trend(
        registrations_total,
        prev_registrations.count(),
    )
    responses_trend_percent, responses_trend_direction = calculate_trend(
        responses_total,
        prev_responses.count(),
    )
    portfolio_trend_percent, portfolio_trend_direction = calculate_trend(
        portfolio_total,
        prev_portfolio.count(),
    )
    active_students_trend_percent, active_students_trend_direction = calculate_trend(
        active_students_total,
        prev_active_students_total,
    )

    activity_total = registrations_total + responses_total + portfolio_total

    def percent(value):
        if not activity_total:
            return 0
        return int(round((value / activity_total) * 100))

    activity_courses_percent = percent(registrations_total)
    activity_vacancies_percent = percent(responses_total)
    activity_portfolio_percent = percent(portfolio_total)

    activity_counts = {'Курсы': registrations_total, 'Вакансии': responses_total, 'Портфолио': portfolio_total}
    main_activity_label = max(activity_counts, key=activity_counts.get) if activity_total else 'пока нет данных'

    students_with_activity = (
        students
        .select_related('study_group')
        .annotate(
            reg_actions=Count(
                'course_registrations',
                filter=Q(
                    course_registrations__status=CourseRegistration.Status.REGISTERED,
                    course_registrations__created_at__gte=period_start,
                    course_registrations__created_at__lt=now,
                ),
            ),
            resp_actions=Count(
                'vacancy_responses',
                filter=Q(
                    vacancy_responses__created_at__gte=period_start,
                    vacancy_responses__created_at__lt=now,
                ),
            ),
            portfolio_actions=Count(
                'portfolio_entries',
                filter=Q(
                    portfolio_entries__created_at__gte=period_start,
                    portfolio_entries__created_at__lt=now,
                ),
            ),
        )
        .annotate(total_actions=F('reg_actions') + F('resp_actions') + F('portfolio_actions'))
        .order_by('-total_actions', 'full_name')
    )
    top_active_students = [student for student in students_with_activity if student.total_actions > 0][:5]
    most_active_student_name = top_active_students[0].full_name if top_active_students else 'пока нет данных'

    students_total = students.count()
    students_without_activity = max(students_total - active_students_total, 0)
    students_with_course_actions = current_registrations.values('student_id').distinct().count()
    students_with_vacancy_actions = current_responses.values('student_id').distinct().count()
    students_without_courses = max(students_total - students_with_course_actions, 0)
    students_without_vacancies = max(students_total - students_with_vacancy_actions, 0)

    context = {
        'selected_period': selected_period,
        'period_options': period_options,
        'include_graduates': include_graduates,

        'registrations_total': registrations_total,
        'responses_total': responses_total,
        'portfolio_total': portfolio_total,
        'active_students_total': active_students_total,

        'registrations_trend_percent': registrations_trend_percent,
        'registrations_trend_direction': registrations_trend_direction,
        'responses_trend_percent': responses_trend_percent,
        'responses_trend_direction': responses_trend_direction,
        'portfolio_trend_percent': portfolio_trend_percent,
        'portfolio_trend_direction': portfolio_trend_direction,
        'active_students_trend_percent': active_students_trend_percent,
        'active_students_trend_direction': active_students_trend_direction,

        'activity_total': activity_total,
        'activity_courses_total': registrations_total,
        'activity_vacancies_total': responses_total,
        'activity_portfolio_total': portfolio_total,
        'activity_courses_percent': activity_courses_percent,
        'activity_vacancies_percent': activity_vacancies_percent,
        'activity_portfolio_percent': activity_portfolio_percent,

        'main_activity_label': main_activity_label,
        'top_active_students': top_active_students,
        'most_active_student_name': most_active_student_name,
        'students_without_activity': students_without_activity,
        'students_without_courses': students_without_courses,
        'students_without_vacancies': students_without_vacancies,
    }
    return render(request, 'curator/activity.html', context)


@role_required(User.Role.CURATOR)
def curator_course_registrations(request):
    students = curator_students_queryset(request.user, include_graduates=True)
    student_ids = students.values_list('id', flat=True)

    registrations = (
        CourseRegistration.objects
        .filter(student_id__in=student_ids)
        .select_related('student', 'student__study_group', 'course')
        .order_by('-created_at')
    )

    q = request.GET.get('q', '').strip()
    course_q = request.GET.get('course', '').strip()
    group_q = request.GET.get('group', '').strip()
    status_q = request.GET.get('status', 'all').strip()

    if q:
        registrations = registrations.filter(
            Q(student__full_name__icontains=q) |
            Q(student__email__icontains=q)
        )
    if course_q:
        registrations = registrations.filter(course__title__icontains=course_q)
    if group_q:
        registrations = registrations.filter(
            Q(student__study_group__name=group_q) |
            Q(student__group__icontains=group_q)
        )
    if status_q in {CourseRegistration.Status.REGISTERED, CourseRegistration.Status.CANCELLED}:
        registrations = registrations.filter(status=status_q)
    else:
        status_q = 'all'

    base_regs = CourseRegistration.objects.filter(student_id__in=student_ids)

    paginator = Paginator(registrations, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(number=page_obj.number, on_each_side=1, on_ends=1)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    context = {
        'registrations': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,

        'q': q,
        'course_q': course_q,
        'group_q': group_q,
        'status_q': status_q,
        'groups': (
            students
            .exclude(study_group__isnull=True)
            .values_list('study_group__name', flat=True)
            .distinct()
            .order_by('study_group__name')
        ),

        'active_count': base_regs.filter(status=CourseRegistration.Status.REGISTERED).count(),
        'cancelled_count': base_regs.filter(status=CourseRegistration.Status.CANCELLED).count(),
        'courses_count': base_regs.values('course_id').distinct().count(),
        'students_count': base_regs.values('student_id').distinct().count(),
    }
    return render(request, 'curator/course_registrations.html', context)


@role_required(User.Role.CURATOR)
def curator_vacancy_responses(request):
    students = curator_students_queryset(request.user, include_graduates=True)
    student_ids = students.values_list('id', flat=True)

    responses = (
        VacancyResponse.objects
        .filter(student_id__in=student_ids)
        .select_related('student', 'student__study_group', 'vacancy')
        .order_by('-created_at')
    )

    q = request.GET.get('q', '').strip()
    vacancy_q = request.GET.get('vacancy', '').strip()
    group_q = request.GET.get('group', '').strip()

    if q:
        responses = responses.filter(
            Q(student__full_name__icontains=q) |
            Q(student__email__icontains=q)
        )
    if vacancy_q:
        responses = responses.filter(vacancy__title__icontains=vacancy_q)
    if group_q:
        responses = responses.filter(
            Q(student__study_group__name=group_q) |
            Q(student__group__icontains=group_q)
        )

    paginator = Paginator(responses, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    rows = []
    for response in page_obj.object_list:
        profile = getattr(response.student, 'student_profile', None)
        resume = getattr(response.student, 'resume_settings', None)
        resume_public_url = ''
        if profile and resume and resume.is_public:
            resume_public_url = request.build_absolute_uri(f"/resumes/public/{profile.public_resume_token}/")
        rows.append((response, resume_public_url))

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    base_resp = VacancyResponse.objects.filter(student_id__in=student_ids)

    context = {
        'rows': rows,
        'page_obj': page_obj,
        'querystring': querystring,

        'q': q,
        'vacancy_q': vacancy_q,
        'group_q': group_q,
        'groups': (
            students
            .exclude(study_group__isnull=True)
            .values_list('study_group__name', flat=True)
            .distinct()
            .order_by('study_group__name')
        ),

        'responses_count': base_resp.count(),
        'students_count': base_resp.values('student_id').distinct().count(),
        'vacancies_count': base_resp.values('vacancy_id').distinct().count(),
    }
    return render(request, 'curator/vacancy_responses.html', context)


@role_required(User.Role.CURATOR)
def curator_students(request):
    include_graduates = request.GET.get('include_graduates') == '1'
    students = (
        curator_students_queryset(request.user, include_graduates=include_graduates)
        .annotate(
            portfolio_total=Count('portfolio_entries'),
            portfolio_pending=Count('portfolio_entries', filter=Q(portfolio_entries__status=PortfolioEntry.Status.PENDING)),
            portfolio_approved=Count('portfolio_entries', filter=Q(portfolio_entries__status=PortfolioEntry.Status.APPROVED)),
            last_activity=Max('portfolio_entries__updated_at'),
        )
        .order_by('full_name')
    )
    return render(request, 'curator/students.html', {'students': students, 'include_graduates': include_graduates})


@role_required(User.Role.CURATOR)
def curator_student_detail(request, student_id):
    student = get_object_or_404(
        curator_students_queryset(request.user, include_graduates=True),
        id=student_id,
    )
    entries = PortfolioEntry.objects.filter(student=student).prefetch_related('attachments').order_by('-created_at')
    resume = getattr(student, 'resume_settings', None)
    profile = getattr(student, 'student_profile', None)
    resume_public_url = ''
    if resume and resume.is_public and profile:
        resume_public_url = request.build_absolute_uri(f"/resumes/public/{profile.public_resume_token}/")

    if request.method == 'POST':
        form = CuratorStudentAcademicStatusForm(request.POST, instance=student)
        if form.is_valid():
            updated_student = form.save(commit=False)
            updated_student.save(update_fields=['academic_status'])
            messages.success(request, 'Учебный статус студента обновлён.')
            return redirect('accounts:curator_student_detail', student_id=student.id)
    else:
        form = CuratorStudentAcademicStatusForm(instance=student)

    context = {
        'student': student,
        'status_form': form,
        'profile': profile,
        'entries': entries,
        'pending_entries': entries.filter(status=PortfolioEntry.Status.PENDING),
        'portfolio_total': entries.count(),
        'portfolio_pending': entries.filter(status=PortfolioEntry.Status.PENDING).count(),
        'portfolio_approved': entries.filter(status=PortfolioEntry.Status.APPROVED).count(),
        'portfolio_rejected': entries.filter(status=PortfolioEntry.Status.REJECTED).count(),
        'resume_public_url': resume_public_url,
    }
    return render(request, 'curator/student_detail.html', context)


@role_required(User.Role.ADMIN)
def admin_dashboard(request):
    students_qs = User.objects.filter(role=User.Role.STUDENT)
    students_total = students_qs.count()
    curators_total = User.objects.filter(role=User.Role.CURATOR).count()

    vacancies = Vacancy.objects.values('status').annotate(total=Count('id'))
    vacancy_summary = {item['status']: item['total'] for item in vacancies}

    courses = Course.objects.values('status').annotate(total=Count('id'))
    course_summary = {item['status']: item['total'] for item in courses}

    specialty_counter = {}
    for student in students_qs.select_related('study_group__specialty_ref'):
        label = (
            getattr(getattr(student.study_group, 'specialty_ref', None), 'name', '')
            or student.specialty
            or 'Без специальности'
        )
        specialty_counter[label] = specialty_counter.get(label, 0) + 1
    specialty_labels = list(specialty_counter.keys())
    specialty_values = list(specialty_counter.values())

    context = {
        'students_total': students_total,
        'students_active': students_qs.filter(is_active=True).count(),
        'students_studying': students_qs.filter(academic_status=User.AcademicStatus.STUDYING).count(),
        'students_academic_leave': students_qs.filter(academic_status=User.AcademicStatus.ACADEMIC_LEAVE).count(),
        'students_graduate': students_qs.filter(academic_status=User.AcademicStatus.GRADUATED).count(),
        'students_inactive_status': students_qs.filter(academic_status=User.AcademicStatus.EXPELLED).count(),
        'inactive_users_total': User.objects.filter(is_active=False).count(),
        'curators_total': curators_total,
        'groups_active': StudyGroup.objects.filter(is_active=True).count(),
        'specialties_total': Specialty.objects.count(),
        'groups_without_curator': StudyGroup.objects.filter(is_active=True, curator__isnull=True).count(),
        'students_without_group': students_qs.filter(study_group__isnull=True).count(),
        'vacancies_active': vacancy_summary.get(Vacancy.Status.ACTIVE, 0),
        'courses_active': course_summary.get(Course.Status.ACTIVE, 0),
        'registrations_total': CourseRegistration.objects.count(),
        'responses_total': VacancyResponse.objects.count(),
        'registrations_registered_total': CourseRegistration.objects.filter(status=CourseRegistration.Status.REGISTERED).count(),
        'registrations_cancelled_total': CourseRegistration.objects.filter(status=CourseRegistration.Status.CANCELLED).count(),
        'activity_portfolio_total': ActivityLog.objects.filter(event_type__startswith='portfolio_').count(),
        'activity_courses_total': ActivityLog.objects.filter(event_type__in=[ActivityLog.EventType.COURSE_REGISTERED, ActivityLog.EventType.COURSE_CANCELLED]).count(),
        'activity_vacancies_total': ActivityLog.objects.filter(event_type=ActivityLog.EventType.VACANCY_APPLIED).count(),
        'latest_activity': ActivityLog.objects.select_related('student').order_by('-created_at')[:10],
        'portfolio_pending_total': PortfolioEntry.objects.filter(status=PortfolioEntry.Status.PENDING).count(),
        'support_new_total': SupportTicket.objects.filter(status=SupportTicket.Status.NEW).count(),
        'support_in_progress_total': SupportTicket.objects.filter(status=SupportTicket.Status.IN_PROGRESS).count(),
        'support_public_total': SupportTicket.objects.filter(source=SupportTicket.Source.PUBLIC).count(),
        'offline_full_courses': Course.objects.filter(format_type=Course.Format.OFFLINE).annotate(
            reg_total=Count(
                'registrations',
                filter=Q(registrations__status=CourseRegistration.Status.REGISTERED),
            )
        ).filter(reg_total__gte=F('places')).count(),
        'vacancy_summary': {
            'active': vacancy_summary.get(Vacancy.Status.ACTIVE, 0),
            'hidden': vacancy_summary.get(Vacancy.Status.HIDDEN, 0),
            'archive': vacancy_summary.get(Vacancy.Status.ARCHIVE, 0),
        },
        'course_summary': {
            'active': course_summary.get(Course.Status.ACTIVE, 0),
            'hidden': course_summary.get(Course.Status.HIDDEN, 0),
            'archive': course_summary.get(Course.Status.ARCHIVE, 0),
        },
        'student_status_chart': [
            students_qs.filter(academic_status=User.AcademicStatus.STUDYING).count(),
            students_qs.filter(academic_status=User.AcademicStatus.ACADEMIC_LEAVE).count(),
            students_qs.filter(academic_status=User.AcademicStatus.GRADUATED).count(),
            students_qs.filter(academic_status=User.AcademicStatus.EXPELLED).count(),
        ],
        'vacancy_status_chart': [
            vacancy_summary.get(Vacancy.Status.ACTIVE, 0),
            vacancy_summary.get(Vacancy.Status.HIDDEN, 0),
            vacancy_summary.get(Vacancy.Status.ARCHIVE, 0),
        ],
        'course_status_chart': [
            course_summary.get(Course.Status.ACTIVE, 0),
            course_summary.get(Course.Status.HIDDEN, 0),
            course_summary.get(Course.Status.ARCHIVE, 0),
        ],
        'specialty_chart_labels': specialty_labels,
        'specialty_chart_values': specialty_values,
        'latest_vacancies': Vacancy.objects.order_by('-created_at')[:5],
        'latest_courses': Course.objects.order_by('-created_at')[:5],
        'latest_students': students_qs.order_by('-date_joined')[:5],
        'latest_admin_actions': AdminActivityLog.objects.select_related('actor').order_by('-created_at')[:8],
    }
    return render(request, 'adminpanel/dashboard.html', context)


@role_required(User.Role.ADMIN)
def admin_activity(request):
    period_options = {
        '30d': {'label': 'За последние 30 дней', 'days': 30},
        '6m': {'label': 'За последние 6 месяцев', 'days': 182},
        '1y': {'label': 'За последний год', 'days': 365},
    }
    selected_period = request.GET.get('period', '30d')
    if selected_period not in period_options:
        selected_period = '30d'

    period_days = period_options[selected_period]['days']
    now = timezone.now()
    period_start = now - timedelta(days=period_days)
    prev_period_start = period_start - timedelta(days=period_days)

    registrations_registered_total = CourseRegistration.objects.filter(
        status=CourseRegistration.Status.REGISTERED,
        created_at__gte=period_start,
        created_at__lt=now,
    ).count()
    responses_total = VacancyResponse.objects.filter(
        created_at__gte=period_start,
        created_at__lt=now,
    ).count()

    prev_registrations_total = CourseRegistration.objects.filter(
        status=CourseRegistration.Status.REGISTERED,
        created_at__gte=prev_period_start,
        created_at__lt=period_start,
    ).count()
    prev_responses_total = VacancyResponse.objects.filter(
        created_at__gte=prev_period_start,
        created_at__lt=period_start,
    ).count()

    activity_portfolio_total = PortfolioEntry.objects.filter(
        created_at__gte=period_start,
        created_at__lt=now,
    ).count()
    activity_courses_total = registrations_registered_total
    activity_vacancies_total = responses_total

    activity_total = activity_courses_total + activity_vacancies_total + activity_portfolio_total

    def percent(value):
        if not activity_total:
            return 0
        return int(round((value / activity_total) * 100))

    def calculate_trend(current, previous):
        if previous == 0:
            if current == 0:
                return 0, 'same'
            return 100, 'up'
        trend_percent = int(round(((current - previous) / previous) * 100))
        if trend_percent > 0:
            return trend_percent, 'up'
        if trend_percent < 0:
            return abs(trend_percent), 'down'
        return 0, 'same'

    registrations_trend_percent, registrations_trend_direction = calculate_trend(
        registrations_registered_total, prev_registrations_total
    )
    responses_trend_percent, responses_trend_direction = calculate_trend(
        responses_total, prev_responses_total
    )

    activity_counts = {
        'Курсы': activity_courses_total,
        'Вакансии': activity_vacancies_total,
        'Портфолио': activity_portfolio_total,
    }
    main_activity_label = max(activity_counts, key=activity_counts.get) if activity_total else 'пока нет данных'

    top_courses_raw = list(
        CourseRegistration.objects
        .filter(
            status=CourseRegistration.Status.REGISTERED,
            created_at__gte=period_start,
            created_at__lt=now,
        )
        .values(title=F('course__title'))
        .annotate(total=Count('id'))
        .order_by('-total', 'title')[:5]
    )
    max_course_total = max([item['total'] for item in top_courses_raw], default=0)
    top_courses = [
        {
            'title': item['title'] or 'Без названия',
            'total': item['total'],
            'percent': round((item['total'] / max_course_total) * 100, 1) if max_course_total else 0,
        }
        for item in top_courses_raw
    ]

    top_vacancies_raw = list(
        VacancyResponse.objects
        .filter(created_at__gte=period_start, created_at__lt=now)
        .values(title=F('vacancy__title'))
        .annotate(total=Count('id'))
        .order_by('-total', 'title')[:5]
    )
    max_vacancy_total = max([item['total'] for item in top_vacancies_raw], default=0)
    top_vacancies = [
        {
            'title': item['title'] or 'Без названия',
            'total': item['total'],
            'percent': round((item['total'] / max_vacancy_total) * 100, 1) if max_vacancy_total else 0,
        }
        for item in top_vacancies_raw
    ]

    context = {
        'registrations_registered_total': registrations_registered_total,
        'responses_total': responses_total,
        'registrations_trend_percent': registrations_trend_percent,
        'responses_trend_percent': responses_trend_percent,
        'registrations_trend_direction': registrations_trend_direction,
        'responses_trend_direction': responses_trend_direction,
        'courses_active': Course.objects.filter(status=Course.Status.ACTIVE).count(),
        'vacancies_active': Vacancy.objects.filter(status=Vacancy.Status.ACTIVE).count(),
        'selected_period': selected_period,
        'period_options': period_options,

        'activity_total': activity_total,
        'activity_portfolio_total': activity_portfolio_total,
        'activity_courses_total': activity_courses_total,
        'activity_vacancies_total': activity_vacancies_total,
        'activity_portfolio_percent': percent(activity_portfolio_total),
        'activity_courses_percent': percent(activity_courses_total),
        'activity_vacancies_percent': percent(activity_vacancies_total),

        'main_activity_label': main_activity_label,
        'top_courses': top_courses,
        'top_vacancies': top_vacancies,
        'top_course_title': top_courses[0]['title'] if top_courses else 'пока нет записей',
        'top_vacancy_title': top_vacancies[0]['title'] if top_vacancies else 'пока нет откликов',
    }
    return render(request, 'adminpanel/activity.html', context)


@role_required(User.Role.ADMIN)
def admin_students(request):
    form = AdminStudentCreateForm()
    if request.method == 'POST':
        form = AdminStudentCreateForm(request.POST)
        if form.is_valid():
            student = form.save()
            log_admin_action(
                request.user,
                AdminActivityLog.Action.CREATE,
                AdminActivityLog.ObjectType.STUDENT,
                student,
                f'Создан студент {student.full_name}.'
            )
            messages.success(request, 'Студент создан.')
            return redirect('accounts:admin_students')

    students = (
        User.objects.filter(role=User.Role.STUDENT)
        .select_related('curator', 'study_group', 'study_group__specialty_ref')
        .annotate(portfolio_total=Count('portfolio_entries'))
        .order_by('full_name')
    )

    q = request.GET.get('q', '').strip()
    group = request.GET.get('group', '').strip()
    curator = request.GET.get('curator', '').strip()
    specialty = request.GET.get('specialty', '').strip()
    academic_status = request.GET.get('academic_status', '').strip()

    if q:
        students = students.filter(Q(full_name__icontains=q) | Q(email__icontains=q))

    if group:
        students = students.filter(Q(study_group__name=group) | Q(group__icontains=group))

    if curator:
        students = students.filter(curator_id=curator)

    if specialty:
        students = students.filter(
            Q(study_group__specialty_ref__name__icontains=specialty) |
            Q(specialty__icontains=specialty)
        )

    if academic_status in {
        User.AcademicStatus.STUDYING,
        User.AcademicStatus.ACADEMIC_LEAVE,
        User.AcademicStatus.GRADUATED,
        User.AcademicStatus.EXPELLED,
    }:
        students = students.filter(academic_status=academic_status)

    paginator = Paginator(students, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(
    number=page_obj.number,
    on_each_side=1,
    on_ends=1,
)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    filter_curators = User.objects.filter(role=User.Role.CURATOR).order_by('full_name')
    filter_groups = StudyGroup.objects.order_by('name')
    filter_specialties = Specialty.objects.order_by('code', 'name')

    return render(
    request,
    'adminpanel/students.html',
    {
        'students': page_obj.object_list,

        # Простые переменные для пагинации, чтобы шаблон не зависел
        # от вложенных обращений page_obj.paginator.*
        'students_page_obj': page_obj,
        'students_page_range': list(page_range),
        'students_current_page': page_obj.number,
        'students_total_pages': page_obj.paginator.num_pages,
        'students_has_pages': page_obj.paginator.num_pages > 1,
        'students_has_previous': page_obj.has_previous(),
        'students_has_next': page_obj.has_next(),
        'students_previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'students_next_page': page_obj.next_page_number() if page_obj.has_next() else None,

        'querystring': querystring,
        'form': form,
        'filter_curators': filter_curators,
        'filter_groups': filter_groups,
        'filter_specialties': filter_specialties,
        'academic_status_filter': academic_status,
    },
)

@role_required(User.Role.ADMIN)
def admin_student_detail(request, student_id):
    student = get_object_or_404(User, id=student_id, role=User.Role.STUDENT)
    form = AdminStudentUpdateForm(instance=student)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            form = AdminStudentUpdateForm(request.POST, instance=student)
            if form.is_valid():
                student = form.save()
                log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.STUDENT, student, f'Обновлён студент {student.full_name}.')
                messages.success(request, 'Данные студента обновлены.')
                return redirect('accounts:admin_student_detail', student_id=student.id)
        elif action == 'reset_password':
            new_password = request.POST.get('temp_password', '').strip()
            if new_password:
                student.set_password(new_password)
                student.must_change_password = True
                student.save(update_fields=['password', 'must_change_password'])
                log_admin_action(request.user, AdminActivityLog.Action.RESET_PASSWORD, AdminActivityLog.ObjectType.STUDENT, student, f'Сброшен пароль студента {student.full_name}.')
                messages.success(request, 'Пароль студента сброшен.')
            else:
                messages.error(request, 'Введите временный пароль.')
            return redirect('accounts:admin_student_detail', student_id=student.id)
        elif action == 'delete':
            confirm_email = request.POST.get('confirm_email', '').strip().lower()
            if confirm_email and confirm_email == (student.email or '').lower():
                student_repr = str(student)
                student.delete()
                log_admin_action(request.user, AdminActivityLog.Action.DELETE, AdminActivityLog.ObjectType.STUDENT, description=f'Удалён студент {student_repr}.')
                messages.success(request, 'Студент удалён.')
                return redirect('accounts:admin_students')
            messages.error(request, 'Email подтверждения не совпадает.')
            return redirect('accounts:admin_student_detail', student_id=student.id)

    group_profile = None
    if student.study_group:
        group_profile = {
            'group_name': student.study_group.name,
            'specialty': student.study_group.specialty_name,
            'admission_year': student.study_group.admission_year,
            'curator': student.study_group.curator,
        }

    portfolio_entries = PortfolioEntry.objects.filter(student=student).prefetch_related('attachments').order_by('-created_at')[:10]

    return render(
        request,
        'adminpanel/student_detail.html',
        {'student': student, 'form': form, 'group_profile': group_profile, 'portfolio_entries': portfolio_entries},
    )


@role_required(User.Role.ADMIN)
def admin_student_import(request):
    report = None
    form = StudentImportForm()
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            created = 0
            skipped = 0
            errors = []
            warnings = []
            rows, parse_errors = parse_import_file(form.cleaned_data['import_file'], 'students')
            if parse_errors:
                messages.error(request, parse_errors[0])
                return redirect('accounts:admin_student_import')

            seen_emails = set()

            for row_idx, row in enumerate(rows, start=2):
                try:
                    if is_empty_import_row(row):
                        skipped += 1
                        warnings.append(f'Строка {row_idx}: пустая строка пропущена.')
                        continue

                    full_name = (row.get('full_name') or '').strip()
                    email = (row.get('email') or '').strip().lower()
                    password = (row.get('password') or '').strip()
                    group_name = (row.get('group') or '').strip()

                    if not full_name:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указано ФИО.')
                        continue
                    if not email:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан email.')
                        continue
                    try:
                        validate_email(email)
                    except ValidationError:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: неверный формат email.')
                        continue
                    if not password:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан временный пароль.')
                        continue
                    if not group_name:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указана группа.')
                        continue
                    if email in seen_emails:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: email {email} повторяется в файле импорта, строка пропущена.')
                        continue
                    if User.objects.filter(email=email).exists():
                        skipped += 1
                        errors.append(f'Строка {row_idx}: пользователь с email {email} уже существует, строка пропущена.')
                        continue

                    normalized_group_name = normalize_group_code(group_name)
                    study_group = StudyGroup.objects.select_related('specialty_ref', 'curator').filter(name=group_name).first()
                    if not study_group and normalized_group_name != group_name:
                        study_group = StudyGroup.objects.select_related('specialty_ref', 'curator').filter(name=normalized_group_name).first()
                    if not study_group:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: группа {group_name} не найдена.')
                        continue

                    with transaction.atomic():
                        student = User(
                            full_name=full_name,
                            email=email,
                            role=User.Role.STUDENT,
                            academic_status=User.AcademicStatus.STUDYING,
                            is_active=True,
                        )
                        student.set_password(password)
                        student.must_change_password = True
                        sync_student_with_group(student, study_group)
                        student.save()
                    created += 1
                    seen_emails.add(email)
                except Exception as exc:
                    skipped += 1
                    errors.append(f'Строка {row_idx}: ошибка обработки: {str(exc)}.')

            report = {'created': created, 'skipped': skipped, 'errors': errors, 'warnings': warnings}
            summary = f'Импорт студентов: создано {created}, пропущено {skipped}, ошибок {len(errors)}.'
            log_admin_action(request.user, AdminActivityLog.Action.CREATE, AdminActivityLog.ObjectType.STUDENT, description=summary)
            messages.success(request, summary)
            if errors:
                messages.error(request, f'При импорте студентов обнаружены ошибки: {len(errors)}.')
            if warnings:
                messages.warning(request, f'При импорте студентов предупреждений: {len(warnings)}.')

    return render(request, 'adminpanel/student_import.html', {'form': form, 'report': report})


@role_required(User.Role.ADMIN)
def admin_curator_import(request):
    report = None
    form = CuratorImportForm()
    if request.method == 'POST':
        form = CuratorImportForm(request.POST, request.FILES)
        if form.is_valid():
            created = 0
            skipped = 0
            errors = []
            warnings = []
            rows, parse_errors = parse_import_file(form.cleaned_data['import_file'], 'curators')
            if parse_errors:
                messages.error(request, parse_errors[0])
                return redirect('accounts:admin_curator_import')

            seen_emails = set()

            for row_idx, row in enumerate(rows, start=2):
                try:
                    if is_empty_import_row(row):
                        skipped += 1
                        warnings.append(f'Строка {row_idx}: пустая строка пропущена.')
                        continue

                    full_name = (row.get('full_name') or '').strip()
                    email = (row.get('email') or '').strip().lower()
                    password = (row.get('password') or '').strip()

                    if not full_name:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указано ФИО.')
                        continue
                    if not email:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан email.')
                        continue
                    try:
                        validate_email(email)
                    except ValidationError:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: неверный формат email.')
                        continue
                    if not password:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан временный пароль.')
                        continue
                    if email in seen_emails:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: email {email} повторяется в файле импорта, строка пропущена.')
                        continue
                    if User.objects.filter(email=email).exists():
                        skipped += 1
                        errors.append(f'Строка {row_idx}: пользователь с email {email} уже существует, строка пропущена.')
                        continue

                    curator = User(full_name=full_name, email=email, role=User.Role.CURATOR, is_active=True, must_change_password=True)
                    curator.set_password(password)
                    curator.save()
                    created += 1
                    seen_emails.add(email)
                except Exception as exc:
                    skipped += 1
                    errors.append(f'Строка {row_idx}: ошибка обработки: {str(exc)}.')

            report = {'created': created, 'skipped': skipped, 'errors': errors, 'warnings': warnings}
            summary = f'Импорт кураторов: создано {created}, пропущено {skipped}, ошибок {len(errors)}.'
            log_admin_action(request.user, AdminActivityLog.Action.CREATE, AdminActivityLog.ObjectType.CURATOR, description=summary)
            messages.success(request, summary)
            if errors:
                messages.error(request, f'При импорте кураторов обнаружены ошибки: {len(errors)}.')
            if warnings:
                messages.warning(request, f'При импорте кураторов предупреждений: {len(warnings)}.')

    return render(request, 'adminpanel/curator_import.html', {'form': form, 'report': report})


@role_required(User.Role.ADMIN)
def admin_group_import(request):
    report = None
    form = GroupImportForm()
    if request.method == 'POST':
        form = GroupImportForm(request.POST, request.FILES)
        if form.is_valid():
            created = 0
            skipped = 0
            errors = []
            warnings = []
            rows, parse_errors = parse_import_file(form.cleaned_data['import_file'], 'groups')
            if parse_errors:
                messages.error(request, parse_errors[0])
                return redirect('accounts:admin_group_import')

            for row_idx, row in enumerate(rows, start=2):
                try:
                    if is_empty_import_row(row):
                        skipped += 1
                        warnings.append(f'Строка {row_idx}: пустая строка пропущена.')
                        continue

                    specialty_letter = (row.get('specialty_letter') or '').strip().upper()
                    admission_year = (row.get('admission_year') or '').strip()
                    course_number = (row.get('course_number') or '').strip()
                    subgroup_number = (row.get('subgroup_number') or '').strip()
                    curator_email = (row.get('curator_email') or '').strip().lower()

                    if not specialty_letter:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан буквенный код специальности.')
                        continue
                    if not admission_year:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан год поступления.')
                        continue
                    if not course_number:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: не указан курс.')
                        continue

                    try:
                        admission_year_int = int(admission_year)
                    except ValueError:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: неверный год поступления.')
                        continue
                    try:
                        course_number_int = int(course_number)
                    except ValueError:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: неверный номер курса.')
                        continue
                    try:
                        subgroup_number_int = int(subgroup_number) if subgroup_number else None
                    except ValueError:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: неверный номер подгруппы.')
                        continue

                    specialty = Specialty.objects.filter(letter_code__iexact=specialty_letter).first()
                    if not specialty:
                        skipped += 1
                        errors.append(f'Строка {row_idx}: специальность {specialty_letter} не найдена, группа не создана.')
                        continue

                    curator = None
                    if curator_email:
                        curator_candidate = User.objects.filter(email=curator_email).first()
                        if not curator_candidate or curator_candidate.role != User.Role.CURATOR:
                            skipped += 1
                            errors.append(f'Строка {row_idx}: куратор {curator_email} не найден, группа не создана.')
                            continue
                        curator = curator_candidate

                    group = StudyGroup(
                        specialty_ref=specialty,
                        admission_year=admission_year_int,
                        course_number=course_number_int,
                        subgroup_number=subgroup_number_int,
                        curator=curator,
                        is_active=True,
                    )
                    group.refresh_name()
                    if StudyGroup.objects.filter(name=group.name).exists():
                        skipped += 1
                        errors.append(f'Строка {row_idx}: группа {group.name} уже существует, строка пропущена.')
                        continue

                    group.save()
                    created += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f'Строка {row_idx}: ошибка обработки: {str(exc)}.')

            report = {'created': created, 'skipped': skipped, 'errors': errors, 'warnings': warnings}
            summary = f'Импорт групп: создано {created}, пропущено {skipped}, ошибок {len(errors)}.'
            log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.GROUP, description=summary)
            messages.success(request, f'Импорт групп завершён: создано {created}, пропущено {skipped}.')
            if errors:
                messages.error(request, f'При импорте групп обнаружены ошибки: {len(errors)}.')
            if warnings:
                messages.warning(request, f'При импорте групп предупреждений: {len(warnings)}.')

    return render(request, 'adminpanel/group_import.html', {'form': form, 'report': report})


@role_required(User.Role.ADMIN)
def admin_import_template_csv(request, import_type):
    if import_type not in IMPORT_TEMPLATES:
        messages.error(request, 'Неизвестный тип шаблона.')
        return redirect('accounts:admin_dashboard')
    return build_template_csv_response(import_type)


@role_required(User.Role.ADMIN)
def admin_import_template_xlsx(request, import_type):
    if import_type not in IMPORT_TEMPLATES:
        messages.error(request, 'Неизвестный тип шаблона.')
        return redirect('accounts:admin_dashboard')
    return build_template_xlsx_response(import_type)


@role_required(User.Role.ADMIN)
def admin_specialties(request):
    edit_id = request.GET.get('edit')
    edit_specialty = None
    if edit_id:
        edit_specialty = get_object_or_404(Specialty, id=edit_id)

    form = AdminSpecialtyForm(instance=edit_specialty)
    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        if action == 'toggle_active':
            specialty = get_object_or_404(Specialty, id=request.POST.get('specialty_id'))
            specialty.is_active = not specialty.is_active
            specialty.save(update_fields=['is_active'])
            messages.success(request, 'Статус специальности обновлён.')
            return redirect('accounts:admin_specialties')

        instance = edit_specialty if action == 'update' else None
        form = AdminSpecialtyForm(request.POST, instance=instance)
        if form.is_valid():
            specialty = form.save()
            action_type = AdminActivityLog.Action.UPDATE if instance else AdminActivityLog.Action.CREATE
            log_admin_action(request.user, action_type, AdminActivityLog.ObjectType.SPECIALTY, specialty, f'Сохранена специальность {specialty.name}.')
            messages.success(request, 'Специальность сохранена.')
            return redirect('accounts:admin_specialties')

    specialties = Specialty.objects.annotate(
        groups_count=Count('study_groups', distinct=True),
        students_count=Count('study_groups__students', distinct=True),
    ).order_by('code', 'name')
    q = request.GET.get('q', '').strip()
    if q:
        specialties = specialties.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(letter_code__icontains=q))

    return render(
        request,
        'adminpanel/specialties.html',
        {'specialties': specialties, 'form': form, 'edit_specialty': edit_specialty},
    )


@role_required(User.Role.ADMIN)
def admin_specialty_detail(request, specialty_id):
    specialty = get_object_or_404(
        Specialty.objects.annotate(
            groups_count=Count('study_groups', distinct=True),
            students_count=Count('study_groups__students', distinct=True),
        ),
        id=specialty_id,
    )
    form = AdminSpecialtyForm(instance=specialty)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_active':
            specialty.is_active = not specialty.is_active
            specialty.save(update_fields=['is_active'])
            messages.success(request, 'Статус специальности обновлён.')
            return redirect('accounts:admin_specialty_detail', specialty_id=specialty.id)

        form = AdminSpecialtyForm(request.POST, instance=specialty)
        if form.is_valid():
            specialty = form.save()
            log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.SPECIALTY, specialty, f'Обновлена специальность {specialty.name}.')
            messages.success(request, 'Данные специальности обновлены.')
            return redirect('accounts:admin_specialty_detail', specialty_id=specialty.id)

    groups = (
        StudyGroup.objects.filter(specialty_ref=specialty)
        .select_related('curator')
        .annotate(students_count=Count('students'))
        .order_by('name')
    )

    return render(
        request,
        'adminpanel/specialty_detail.html',
        {'specialty': specialty, 'form': form, 'groups': groups},
    )


@role_required(User.Role.ADMIN)
def admin_academic_structure(request):
    specialties = Specialty.objects.order_by('code', 'name')[:20]
    groups = StudyGroup.objects.select_related('specialty_ref', 'curator').annotate(
        students_count=Count('students')
    ).order_by('name')[:20]
    return render(
        request,
        'adminpanel/academic_structure.html',
        {'specialties': specialties, 'groups': groups},
    )


@role_required(User.Role.ADMIN)
def admin_groups(request):
    form = AdminStudyGroupForm()

    if request.method == 'POST':
        action = request.POST.get('action', 'create')

        if action == 'toggle_active':
            group = get_object_or_404(StudyGroup, id=request.POST.get('group_id'))
            group.is_active = not group.is_active
            group.save(update_fields=['is_active'])
            if group.is_active:
                sync_group_students(group)
            messages.success(request, 'Статус группы обновлён.')
            return redirect('accounts:admin_groups')

        form = AdminStudyGroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)
            group.refresh_name()
            group.save()
            sync_group_students(group)

            log_admin_action(
                request.user,
                AdminActivityLog.Action.CREATE,
                AdminActivityLog.ObjectType.GROUP,
                group,
                f'Создана группа {group.name}.'
            )

            messages.success(request, 'Группа сохранена.')
            return redirect('accounts:admin_groups')

    groups = (
        StudyGroup.objects
        .select_related('specialty_ref', 'curator')
        .annotate(students_count=Count('students'))
        .order_by('name')
    )

    q = request.GET.get('q', '').strip()
    specialty = request.GET.get('specialty', '').strip()
    curator = request.GET.get('curator', '').strip()
    course_number = request.GET.get('course_number', '').strip()
    admission_year = request.GET.get('admission_year', '').strip()

    if q:
        groups = groups.filter(name__icontains=q)

    if specialty:
        groups = groups.filter(specialty_ref_id=specialty)

    if curator:
        groups = groups.filter(curator_id=curator)

    if course_number:
        groups = groups.filter(course_number=course_number)

    if admission_year:
        groups = groups.filter(admission_year=admission_year)

    paginator = Paginator(groups, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(number=page_obj.number, on_each_side=1, on_ends=1)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    return render(
        request,
        'adminpanel/groups.html',
        {
            'groups': page_obj.object_list,
            'page_obj': page_obj,
            'form': form,
            'specialties': Specialty.objects.order_by('code', 'name'),
            'curators': User.objects.filter(role=User.Role.CURATOR).order_by('full_name'),

            'querystring': querystring,
            'specialty_filter': specialty,
            'curator_filter': curator,
            'course_number_filter': course_number,
            'admission_year_filter': admission_year,

            'groups_page_range': list(page_range),
            'groups_current_page': page_obj.number,
            'groups_has_pages': page_obj.paginator.num_pages > 1,
            'groups_has_previous': page_obj.has_previous(),
            'groups_has_next': page_obj.has_next(),
            'groups_previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'groups_next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        },
    )


@role_required(User.Role.ADMIN)
def admin_group_detail(request, group_id):
    group = get_object_or_404(StudyGroup.objects.select_related('specialty_ref', 'curator'), id=group_id)
    form = AdminStudyGroupForm(instance=group)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update':
            old_sync_signature = (
                group.name, group.specialty_ref_id, group.admission_year, group.course_number, group.subgroup_number, group.curator_id
            )
            form = AdminStudyGroupForm(request.POST, instance=group)
            if form.is_valid():
                group = form.save(commit=False)
                group.refresh_name()
                group.save()
                new_sync_signature = (
                    group.name, group.specialty_ref_id, group.admission_year, group.course_number, group.subgroup_number, group.curator_id
                )
                if old_sync_signature != new_sync_signature:
                    sync_group_students(group)
                log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.GROUP, group, f'Обновлена группа {group.name}.')
                messages.success(request, 'Данные группы обновлены.')
                return redirect('accounts:admin_group_detail', group_id=group.id)
        elif action == 'toggle_active':
            group.is_active = not group.is_active
            group.save(update_fields=['is_active'])
            messages.success(request, 'Статус группы обновлён.')
            return redirect('accounts:admin_group_detail', group_id=group.id)
        elif action == 'sync_students':
            sync_group_students(group)
            messages.success(request, 'Данные студентов группы синхронизированы.')
            return redirect('accounts:admin_group_detail', group_id=group.id)

    students = User.objects.filter(role=User.Role.STUDENT, study_group=group).order_by('full_name')
    return render(
        request,
        'adminpanel/group_detail.html',
        {'group': group, 'form': form, 'students': students, 'students_count': students.count()},
    )


@role_required(User.Role.ADMIN)
def admin_curators(request):
    form = AdminCuratorCreateForm()

    if request.method == 'POST':
        form = AdminCuratorCreateForm(request.POST)
        if form.is_valid():
            curator = form.save()
            log_admin_action(
                request.user,
                AdminActivityLog.Action.CREATE,
                AdminActivityLog.ObjectType.CURATOR,
                curator,
                f'Создан куратор {curator.full_name}.'
            )
            messages.success(request, 'Куратор создан.')
            return redirect('accounts:admin_curators')

    q = request.GET.get('q', '').strip()
    group = request.GET.get('group', '').strip()
    student = request.GET.get('student', '').strip()

    curators = (
        User.objects.filter(role=User.Role.CURATOR)
        .annotate(
            groups_count=Count('managed_study_groups', distinct=True),
            students_count=Count('managed_study_groups__students', distinct=True),
        )
        .order_by('full_name')
    )

    if q:
        curators = curators.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q)
        )

    if group.isdigit():
        curators = curators.filter(managed_study_groups__id=int(group))

    if student.isdigit():
        curators = curators.filter(managed_study_groups__students__id=int(student))

    curators = curators.distinct()

    paginator = Paginator(curators, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(number=page_obj.number, on_each_side=1, on_ends=1)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    filter_groups = StudyGroup.objects.order_by('name')
    filter_students = (
        User.objects.filter(role=User.Role.STUDENT)
        .select_related('study_group')
        .order_by('full_name')
    )

    return render(
        request,
        'adminpanel/curators.html',
        {
            'curators': page_obj.object_list,
            'page_obj': page_obj,
            'form': form,
            'querystring': querystring,

            'filter_groups': filter_groups,
            'filter_students': filter_students,
            'group_filter': group,
            'student_filter': student,

            'curators_page_range': list(page_range),
            'curators_current_page': page_obj.number,
            'curators_has_pages': page_obj.paginator.num_pages > 1,
            'curators_has_previous': page_obj.has_previous(),
            'curators_has_next': page_obj.has_next(),
            'curators_previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'curators_next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        },
    )


@role_required(User.Role.ADMIN)
def admin_curator_detail(request, curator_id):
    curator = get_object_or_404(User, id=curator_id, role=User.Role.CURATOR)
    form = AdminCuratorUpdateForm(instance=curator)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            form = AdminCuratorUpdateForm(request.POST, instance=curator)
            if form.is_valid():
                curator = form.save()
                log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.CURATOR, curator, f'Обновлён куратор {curator.full_name}.')
                messages.success(request, 'Данные куратора обновлены.')
                return redirect('accounts:admin_curator_detail', curator_id=curator.id)
        elif action == 'toggle_active':
            curator.is_active = not curator.is_active
            curator.save(update_fields=['is_active'])
            messages.success(request, 'Статус куратора обновлён.')
            return redirect('accounts:admin_curator_detail', curator_id=curator.id)
        elif action == 'reset_password':
            new_password = request.POST.get('temp_password', '').strip()
            if new_password:
                curator.set_password(new_password)
                curator.must_change_password = True
                curator.save(update_fields=['password', 'must_change_password'])
                log_admin_action(request.user, AdminActivityLog.Action.RESET_PASSWORD, AdminActivityLog.ObjectType.CURATOR, curator, f'Сброшен пароль куратора {curator.full_name}.')
                messages.success(request, 'Пароль куратора сброшен.')
            else:
                messages.error(request, 'Введите временный пароль.')
            return redirect('accounts:admin_curator_detail', curator_id=curator.id)
        elif action == 'delete':
            confirm_email = request.POST.get('confirm_email', '').strip().lower()
            if confirm_email and confirm_email == (curator.email or '').lower():
                curator.delete()
                messages.success(request, 'Куратор удалён.')
                return redirect('accounts:admin_curators')
            messages.error(request, 'Email подтверждения не совпадает.')
            return redirect('accounts:admin_curator_detail', curator_id=curator.id)

    groups = StudyGroup.objects.filter(curator=curator).order_by('name')
    students = (
        User.objects.filter(role=User.Role.STUDENT)
        .filter(Q(study_group__curator=curator) | Q(curator=curator))
        .distinct()
        .order_by('full_name')
    )
    return render(
        request,
        'adminpanel/curator_detail.html',
        {'curator': curator, 'form': form, 'students': students, 'groups': groups},
    )


@role_required(User.Role.ADMIN)
def admin_vacancies(request):
    status_filter = request.GET.get('status', 'all')
    q = request.GET.get('q', '').strip()
    format_filter = request.GET.get('format_type', '').strip()
    employment_filter = request.GET.get('employment_type', '').strip()
    direction_filter = request.GET.get('direction', '').strip()

    form = AdminVacancyForm()
    if request.method == 'POST':
        form = AdminVacancyForm(request.POST)
        if form.is_valid():
            vacancy = form.save()
            log_admin_action(
                request.user,
                AdminActivityLog.Action.CREATE,
                AdminActivityLog.ObjectType.VACANCY,
                vacancy,
                f'Создана вакансия {vacancy.title}.'
            )
            messages.success(request, 'Вакансия создана.')
            return redirect('accounts:admin_vacancies')

    vacancies = Vacancy.objects.order_by('-created_at')

    filter_directions = (
        Vacancy.objects
        .exclude(direction='')
        .exclude(direction__isnull=True)
        .values_list('direction', flat=True)
        .distinct()
        .order_by('direction')
    )

    if q:
        vacancies = vacancies.filter(
            Q(title__icontains=q) |
            Q(company__icontains=q)
        )

    if status_filter in {Vacancy.Status.ACTIVE, Vacancy.Status.HIDDEN, Vacancy.Status.ARCHIVE}:
        vacancies = vacancies.filter(status=status_filter)
    else:
        status_filter = 'all'

    if format_filter:
        vacancies = vacancies.filter(format_type__icontains=format_filter)

    if employment_filter:
        vacancies = vacancies.filter(employment_type__icontains=employment_filter)

    if direction_filter:
        vacancies = vacancies.filter(direction__icontains=direction_filter)

    paginator = Paginator(vacancies, 8)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(number=page_obj.number, on_each_side=1, on_ends=1)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    return render(
        request,
        'adminpanel/vacancies.html',
        {
            'vacancies': page_obj.object_list,
            'page_obj': page_obj,
            'page_range': page_range,
            'querystring': querystring,
            'form': form,
            'status_filter': status_filter,
            'filter_directions': filter_directions,
        },
    )


@role_required(User.Role.ADMIN)
def admin_vacancy_detail(request, vacancy_id):
    vacancy = get_object_or_404(Vacancy, id=vacancy_id)
    form = AdminVacancyForm(instance=vacancy)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update':
            form = AdminVacancyForm(request.POST, instance=vacancy)
            if form.is_valid():
                vacancy = form.save()
                log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.VACANCY, vacancy, f'Обновлена вакансия {vacancy.title}.')
                messages.success(request, 'Вакансия обновлена.')
                return redirect('accounts:admin_vacancy_detail', vacancy_id=vacancy.id)
        elif action == 'set_status':
            new_status = request.POST.get('status')
            if new_status in {Vacancy.Status.ACTIVE, Vacancy.Status.HIDDEN, Vacancy.Status.ARCHIVE}:
                vacancy.status = new_status
                vacancy.save(update_fields=['status', 'updated_at'])
                log_admin_action(request.user, AdminActivityLog.Action.STATUS_CHANGE, AdminActivityLog.ObjectType.VACANCY, vacancy, f'Статус вакансии изменён на "{vacancy.get_status_display()}".')
                messages.success(request, 'Статус вакансии обновлён.')
                return redirect('accounts:admin_vacancy_detail', vacancy_id=vacancy.id)
        elif action == 'delete':
            confirm_title = request.POST.get('confirm_title', '').strip()
            if confirm_title == vacancy.title:
                vacancy.delete()
                messages.success(request, 'Вакансия удалена.')
                return redirect('accounts:admin_vacancies')
            messages.error(request, 'Название вакансии для подтверждения введено неверно.')
            return redirect('accounts:admin_vacancy_detail', vacancy_id=vacancy.id)

    return render(request, 'adminpanel/vacancy_detail.html', {'vacancy': vacancy, 'form': form})


@role_required(User.Role.ADMIN)
def admin_courses(request):
    status_filter = request.GET.get('status', 'all')
    q = request.GET.get('q', '').strip()
    kind_filter = request.GET.get('kind', '').strip()
    format_filter = request.GET.get('format_type', '').strip()

    form = AdminCourseForm()
    if request.method == 'POST':
        form = AdminCourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            log_admin_action(
                request.user,
                AdminActivityLog.Action.CREATE,
                AdminActivityLog.ObjectType.COURSE,
                course,
                f'Создан курс {course.title}.'
            )
            messages.success(request, 'Курс создан.')
            return redirect('accounts:admin_courses')

    courses = Course.objects.annotate(
        registrations_count=Count(
            'registrations',
            filter=Q(registrations__status=CourseRegistration.Status.REGISTERED),
        )
    ).order_by('-created_at')

    if q:
        courses = courses.filter(
            Q(title__icontains=q) |
            Q(organization__icontains=q)
        )

    if status_filter in {Course.Status.ACTIVE, Course.Status.HIDDEN, Course.Status.ARCHIVE}:
        courses = courses.filter(status=status_filter)
    else:
        status_filter = 'all'

    if kind_filter in {Course.Kind.COURSE, Course.Kind.SEMINAR, Course.Kind.PRACTICE}:
        courses = courses.filter(kind=kind_filter)

    if format_filter in {Course.Format.ONLINE, Course.Format.OFFLINE}:
        courses = courses.filter(format_type=format_filter)

    paginator = Paginator(courses, 8)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_range = paginator.get_elided_page_range(number=page_obj.number, on_each_side=1, on_ends=1)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    return render(
        request,
        'adminpanel/courses.html',
        {
            'courses': page_obj.object_list,
            'page_obj': page_obj,
            'page_range': page_range,
            'querystring': querystring,
            'form': form,
            'status_filter': status_filter,
        },
    )


@role_required(User.Role.ADMIN)
def admin_course_detail(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    form = AdminCourseForm(instance=course)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update':
            form = AdminCourseForm(request.POST, instance=course)
            if form.is_valid():
                course = form.save()
                log_admin_action(request.user, AdminActivityLog.Action.UPDATE, AdminActivityLog.ObjectType.COURSE, course, f'Обновлён курс {course.title}.')
                messages.success(request, 'Курс обновлён.')
                return redirect('accounts:admin_course_detail', course_id=course.id)
        elif action == 'set_status':
            new_status = request.POST.get('status')
            if new_status in {Course.Status.ACTIVE, Course.Status.HIDDEN, Course.Status.ARCHIVE}:
                course.status = new_status
                course.save(update_fields=['status', 'updated_at'])
                log_admin_action(request.user, AdminActivityLog.Action.STATUS_CHANGE, AdminActivityLog.ObjectType.COURSE, course, f'Статус курса изменён на "{course.get_status_display()}".')
                messages.success(request, 'Статус курса обновлён.')
                return redirect('accounts:admin_course_detail', course_id=course.id)
        elif action == 'delete':
            confirm_title = request.POST.get('confirm_title', '').strip()
            if confirm_title == course.title:
                course.delete()
                messages.success(request, 'Курс удалён.')
                return redirect('accounts:admin_courses')
            messages.error(request, 'Название курса для подтверждения введено неверно.')
            return redirect('accounts:admin_course_detail', course_id=course.id)

    registrations_qs = CourseRegistration.objects.filter(course=course).select_related('student', 'student__study_group')
    active_registrations_count = registrations_qs.filter(status=CourseRegistration.Status.REGISTERED).count()
    cancelled_registrations_count = registrations_qs.filter(status=CourseRegistration.Status.CANCELLED).count()
    return render(
        request,
        'adminpanel/course_detail.html',
        {
            'course': course,
            'form': form,
            'registrations_count': active_registrations_count,
            'course_registrations': registrations_qs,
            'active_registrations_count': active_registrations_count,
            'cancelled_registrations_count': cancelled_registrations_count,
            'free_places_count': max(course.places - active_registrations_count, 0),
        },
    )


@role_required(User.Role.ADMIN)
def admin_responses(request):
    responses = (
        VacancyResponse.objects
        .select_related('student', 'student__study_group', 'vacancy')
        .order_by('-created_at')
    )

    q = request.GET.get('q', '').strip()
    vacancy = request.GET.get('vacancy', '').strip()
    group = request.GET.get('group', '').strip()
    specialty = request.GET.get('specialty', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if q:
        responses = responses.filter(
            Q(student__full_name__icontains=q) |
            Q(student__email__icontains=q)
        )
    if vacancy:
        responses = responses.filter(vacancy__title__icontains=vacancy)
    if group:
        responses = responses.filter(
            Q(student__study_group__name=group) |
            Q(student__group__icontains=group)
        )
    if specialty:
        responses = responses.filter(
            Q(student__study_group__specialty_ref__name__icontains=specialty) |
            Q(student__specialty__icontains=specialty)
        )
    if date_from:
        responses = responses.filter(created_at__date__gte=date_from)
    if date_to:
        responses = responses.filter(created_at__date__lte=date_to)

    paginator = Paginator(responses, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    rows = []
    for response in page_obj.object_list:
        resume_url = response.resume_link_snapshot or ''
        rows.append((response, resume_url))

    base_responses = VacancyResponse.objects.select_related('student', 'vacancy')

    context = {
        'rows': rows,
        'page_obj': page_obj,
        'querystring': querystring,

        'q': q,
        'vacancy_q': vacancy,
        'group_q': group,
        'specialty_q': specialty,
        'date_from': date_from,
        'date_to': date_to,

        'responses_count': base_responses.count(),
        'students_count': base_responses.values('student_id').distinct().count(),
        'vacancies_count': base_responses.values('vacancy_id').distinct().count(),
        'groups_count': (
            User.objects
            .filter(role=User.Role.STUDENT, vacancy_responses__isnull=False)
            .values('study_group_id', 'group')
            .distinct()
            .count()
        ),
    }
    return render(request, 'adminpanel/responses.html', context)


@role_required(User.Role.ADMIN)
def admin_course_registrations(request):
    registrations = (
        CourseRegistration.objects
        .select_related('student', 'student__study_group', 'course')
        .order_by('-created_at')
    )

    q = request.GET.get('q', '').strip()
    course = request.GET.get('course', '').strip()
    group = request.GET.get('group', '').strip()
    format_type = request.GET.get('format_type', '').strip()
    status = request.GET.get('status', 'all').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    if q:
        registrations = registrations.filter(
            Q(student__full_name__icontains=q) |
            Q(student__email__icontains=q)
        )
    if course:
        registrations = registrations.filter(course__title__icontains=course)
    if group:
        registrations = registrations.filter(
            Q(student__study_group__name=group) |
            Q(student__group__icontains=group)
        )
    if format_type in {Course.Format.ONLINE, Course.Format.OFFLINE}:
        registrations = registrations.filter(course__format_type=format_type)
    else:
        format_type = ''
    if status in {CourseRegistration.Status.REGISTERED, CourseRegistration.Status.CANCELLED}:
        registrations = registrations.filter(status=status)
    else:
        status = 'all'
    if date_from:
        registrations = registrations.filter(created_at__date__gte=date_from)
    if date_to:
        registrations = registrations.filter(created_at__date__lte=date_to)

    paginator = Paginator(registrations, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    base_regs = CourseRegistration.objects.select_related('student', 'course')

    context = {
        'registrations': page_obj.object_list,
        'page_obj': page_obj,
        'querystring': querystring,

        'q': q,
        'course_q': course,
        'group_q': group,
        'format_type': format_type,
        'status_q': status,
        'date_from': date_from,
        'date_to': date_to,

        'active_count': base_regs.filter(status=CourseRegistration.Status.REGISTERED).count(),
        'cancelled_count': base_regs.filter(status=CourseRegistration.Status.CANCELLED).count(),
        'courses_count': base_regs.values('course_id').distinct().count(),
        'students_count': base_regs.values('student_id').distinct().count(),
    }
    return render(request, 'adminpanel/course_registrations.html', context)


@role_required(User.Role.ADMIN)
def admin_support_tickets(request):
    tickets = (
        SupportTicket.objects
        .select_related('student', 'student__study_group', 'requester', 'requester__study_group')
        .order_by('-created_at')
    )

    status = request.GET.get('status', '').strip()
    category = request.GET.get('category', '').strip()
    source = request.GET.get('source', '').strip()
    requester_type = request.GET.get('requester_type', '').strip()

    if status in {choice[0] for choice in SupportTicket.Status.choices}:
        tickets = tickets.filter(status=status)
    if category in {choice[0] for choice in SupportTicket.Category.choices}:
        tickets = tickets.filter(category=category)
    if source in {choice[0] for choice in SupportTicket.Source.choices}:
        tickets = tickets.filter(source=source)
    if requester_type in {choice[0] for choice in SupportTicket.RequesterType.choices}:
        tickets = tickets.filter(requester_type=requester_type)

    base_tickets = SupportTicket.objects.all()

    def make_initials(name):
        cleaned = (name or '').strip()
        if not cleaned:
            return '?'
        parts = [part for part in cleaned.replace('-', ' ').split() if part]
        if len(parts) >= 2:
            return f'{parts[0][0]}{parts[1][0]}'.upper()
        return cleaned[:2].upper()

    ticket_rows = []
    for ticket in tickets:
        requester = ticket.requester or ticket.student

        avatar_url = ''
        if requester and requester.photo:
            try:
                avatar_url = requester.photo.url
            except ValueError:
                avatar_url = ''

        if requester:
            author_name = requester.full_name or requester.email
            author_role = requester.get_role_display()
            author_group = ''
            if requester.study_group:
                author_group = requester.study_group.name
            elif requester.group:
                author_group = requester.group
        else:
            author_name = ticket.public_full_name or 'Пользователь без входа'
            author_role = 'Без входа'
            author_group = ticket.public_email or ticket.public_contact or ''

        ticket_rows.append({
            'ticket': ticket,
            'avatar_url': avatar_url,
            'initials': make_initials(author_name),
            'author_name': author_name,
            'author_role': author_role,
            'author_group': author_group,
        })

    status_tabs = [
        {
            'value': '',
            'label': 'Все',
            'count': base_tickets.count(),
            'icon': 'bi-list-ul',
            'class': 'is-all',
        },
        {
            'value': SupportTicket.Status.NEW,
            'label': 'Новые',
            'count': base_tickets.filter(status=SupportTicket.Status.NEW).count(),
            'icon': 'bi-dot',
            'class': 'is-new',
        },
        {
            'value': SupportTicket.Status.IN_PROGRESS,
            'label': 'В работе',
            'count': base_tickets.filter(status=SupportTicket.Status.IN_PROGRESS).count(),
            'icon': 'bi-dot',
            'class': 'is-progress',
        },
        {
            'value': SupportTicket.Status.RESOLVED,
            'label': 'Решено',
            'count': base_tickets.filter(status=SupportTicket.Status.RESOLVED).count(),
            'icon': 'bi-dot',
            'class': 'is-resolved',
        },
        {
            'value': SupportTicket.Status.CLOSED,
            'label': 'Закрыто',
            'count': base_tickets.filter(status=SupportTicket.Status.CLOSED).count(),
            'icon': 'bi-archive',
            'class': 'is-closed',
        },
    ]

    query_params = request.GET.copy()
    query_params.pop('status', None)
    query_params.pop('page', None)
    status_querystring = query_params.urlencode()

    return render(
        request,
        'adminpanel/support_tickets.html',
        {
            'tickets': tickets,
            'ticket_rows': ticket_rows,
            'status_filter': status,
            'category_filter': category,
            'status_choices': SupportTicket.Status.choices,
            'category_choices': SupportTicket.Category.choices,
            'source_choices': SupportTicket.Source.choices,
            'requester_type_choices': SupportTicket.RequesterType.choices,
            'source_filter': source,
            'requester_type_filter': requester_type,
            'status_tabs': status_tabs,
            'status_querystring': status_querystring,
        },
    )


@role_required(User.Role.ADMIN)
def admin_support_ticket_detail(request, ticket_id):
    ticket = get_object_or_404(SupportTicket.objects.select_related('student', 'student__study_group', 'requester'), id=ticket_id)
    form = SupportTicketAdminUpdateForm(instance=ticket)
    if request.method == 'POST':
        form = SupportTicketAdminUpdateForm(request.POST, instance=ticket)
        if form.is_valid():
            updated_ticket = form.save(commit=False)
            old_status = ticket.status
            if updated_ticket.status in {SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED}:
                updated_ticket.resolved_at = updated_ticket.resolved_at or timezone.now()
            else:
                updated_ticket.resolved_at = None
            updated_ticket.save()
            if old_status != updated_ticket.status:
                log_admin_action(
                    request.user,
                    AdminActivityLog.Action.STATUS_CHANGE,
                    AdminActivityLog.ObjectType.SUPPORT_TICKET,
                    updated_ticket,
                    f'Статус обращения изменён: {ticket.get_status_display()} → {updated_ticket.get_status_display()}.',
                )
            messages.success(request, 'Обращение обновлено.')
            return redirect('accounts:admin_support_ticket_detail', ticket_id=ticket.id)
    
    return render(request, 'adminpanel/support_ticket_detail.html', {'ticket': ticket, 'form': form})


@role_required(User.Role.ADMIN)
def admin_activity_log(request):
    logs = AdminActivityLog.objects.select_related('actor').order_by('-created_at')
    q = request.GET.get('q', '').strip()
    object_type = request.GET.get('object_type', '').strip()
    action = request.GET.get('action', '').strip()
    actor = request.GET.get('actor', '').strip()

    if q:
        logs = logs.filter(
            Q(actor__email__icontains=q) |
            Q(actor__full_name__icontains=q) |
            Q(object_repr__icontains=q) |
            Q(description__icontains=q)
        )
    if object_type in {choice[0] for choice in AdminActivityLog.ObjectType.choices}:
        logs = logs.filter(object_type=object_type)
    if action in {choice[0] for choice in AdminActivityLog.Action.choices}:
        logs = logs.filter(action=action)
    if actor.isdigit():
        logs = logs.filter(actor_id=int(actor))

    paginator = Paginator(logs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(
        request,
        'adminpanel/activity_log.html',
        {
            'page_obj': page_obj,
            'logs': page_obj.object_list,
            'q': q,
            'object_type_filter': object_type,
            'action_filter': action,
            'actor_filter': actor,
            'object_type_choices': AdminActivityLog.ObjectType.choices,
            'action_choices': AdminActivityLog.Action.choices,
            'admin_users': User.objects.filter(role=User.Role.ADMIN).order_by('full_name'),
        },
    )


@role_required(User.Role.STUDENT)
def profile_edit(request):
    profile = getattr(request.user, 'student_profile', None)
    if profile is None:
        from .models import StudentProfile
        profile = StudentProfile.objects.create(user=request.user)

    if request.method == 'POST' and request.POST.get('action') == 'remove_photo':
        apply_user_photo_update(request, request.user)
        messages.success(request, 'Фото удалено.')
        return redirect('accounts:profile_edit')

    if request.method == 'POST':
        user_form = UserStudentForm(request.POST, request.FILES, instance=request.user)
        profile_form = StudentProfileForm(request.POST, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            previous_photo = request.user.photo if request.user.photo else None
            user = user_form.save()
            apply_user_photo_update(request, user)
            profile_form.save()
            return redirect('accounts:profile_edit')
    else:
        user_form = UserStudentForm(instance=request.user)
        profile_form = StudentProfileForm(instance=profile)

    academic_form = StudentAcademicReadonlyForm(instance=request.user)
    resume = getattr(request.user, 'resume_settings', None)
    resume_public_url = request.build_absolute_uri(f"/resumes/public/{profile.public_resume_token}/") if profile else ''
    if request.user.study_group and request.user.study_group.course_number:
        current_course = request.user.study_group.course_number
    elif request.user.admission_year:
        current_course = max(date.today().year - request.user.admission_year + 1, 1)
    else:
        current_course = None
    return render(request, 'accounts/profile_edit.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'academic_form': academic_form,
        'resume': resume,
        'resume_public_url': resume_public_url,
        'current_course': current_course,
    })


@role_required(User.Role.CURATOR)
def curator_profile_edit(request):
    if request.method == 'POST' and request.POST.get('action') == 'remove_photo':
        apply_user_photo_update(request, request.user)
        messages.success(request, 'Фото удалено.')
        return redirect('accounts:curator_profile_edit')

    if request.method == 'POST':
        user_form = UserProfileSettingsForm(request.POST, request.FILES, instance=request.user)
        if user_form.is_valid():
            user = user_form.save()
            apply_user_photo_update(request, user)
            messages.success(request, 'Профиль куратора сохранён.')
            return redirect('accounts:curator_profile_edit')
    else:
        user_form = UserProfileSettingsForm(instance=request.user)

    return render(request, 'accounts/user_profile_edit.html', {'user_form': user_form, 'profile_title': 'Настройки куратора'})


@role_required(User.Role.ADMIN)
def admin_profile_edit(request):
    if request.method == 'POST':
        user_form = UserProfileSettingsForm(request.POST, request.FILES, instance=request.user)
        if user_form.is_valid():
            user = user_form.save()
            apply_user_photo_update(request, user)
            return redirect('accounts:admin_profile_edit')
    else:
        user_form = UserProfileSettingsForm(instance=request.user)

    return render(request, 'accounts/user_profile_edit.html', {'user_form': user_form, 'profile_title': 'Настройки администратора'})
